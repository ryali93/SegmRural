# -*-coding: utf-8-*-

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl import drawing
from Generales import settings
import win32com.client
import os
import string
import xlsxwriter

conn = settings.conectionDB_pymmsql()
# conn2 = settings.conectionGDB_pymmsql()

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

titleFill = PatternFill(start_color='D9D9D9',
                        end_color='D9D9D9',
                        fill_type='solid')

def abc(inicial, final):
    abc = list(string.ascii_uppercase)
    ini = abc.index(inicial)
    fin = abc.index(final)
    lista = abc[ini:fin + 1]
    return lista

def actualizaIDEMP(ubigeo):
    cursor = conn.cursor()
    cursor.execute("UPDATE CPV_SEGMENTACION.dbo.SEGM_R_EMP SET IDEMP = IDRUTA + EMP WHERE UBIGEO = '{}'".format(ubigeo))
    conn.commit()
    cursor.close()

def actualizarAsignRural(ubigeo):
    cursor = conn.cursor()
    cursor.execute("EXEC USP_ACTUALIZA_ASIGN_RURAL '{}'".format(ubigeo))
    conn.commit()
    cursor.close()

def infoEmpadronador(ubigeo, conn=conn):
    cursor = conn.cursor()
    cursor.execute("SELECT EMP, IDEMP, SCR, IDRUTA, IDSCR FROM CPV_SEGMENTACION.dbo.SEGM_R_EMP WHERE UBIGEO = '{}' ORDER BY EMP".format(ubigeo))
    informacion = [[x[0], x[1], x[2], x[3], x[4]] for x in cursor]
    cursor.close()
    return informacion

def infoEmpadronadorSCR(ubigeo, conn=conn):
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT IDSCR, EMP FROM CPV_SEGMENTACION.dbo.SEGM_R_EMP WHERE UBIGEO = '{}'".format(ubigeo))
    informacion = [[x[0], x[1]] for x in cursor]
    cursor.close()
    return informacion

def infoSCR(ubigeo, conn=conn):
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT IDSCR FROM CPV_SEGMENTACION.dbo.SEGM_R_EMP WHERE UBIGEO = '{}'".format(ubigeo))
    informacion = [x[0] for x in cursor]
    cursor.close()
    return informacion

def infoGeneral(ubigeo, conn=conn):
    cursor = conn.cursor()
    cursor.execute("SELECT DEPARTAMENTO, PROVINCIA, DISTRITO FROM CPV_SEGMENTACION.dbo.VW_DISTRITO WHERE UBIGEO = '{}'".format(ubigeo))
    informacion = [[ubigeo[0:2], ubigeo[2:4], ubigeo[4:6], x[0], x[1], x[2]] for x in cursor]
    cursor.close()
    return informacion[0]

def set_border(cell_range, ws):
    lista = abc('A', 'Q')
    row = cell_range.split(":")[0][1]
    ini = cell_range.split(":")[0][0]
    fin = cell_range.split(":")[1][0]
    column = range(lista.index(ini) + 1, lista.index(fin) + 2)
    for cell in column:
        ws.cell(row=int(row), column=cell).border = thin_border

def cuerpo_border(cell_range, ws):
    lista = abc('A', 'Q')
    rowini = cell_range.split(":")[0][1:3]
    rowfin = cell_range.split(":")[1][1:3]
    ini = cell_range.split(":")[0][0]
    fin = cell_range.split(":")[1][0]
    rows = range(int(rowini), int(rowfin) + 1)
    column = range(lista.index(ini) + 1, lista.index(fin) + 2)
    for x in lista:
        ws.column_dimensions[x].width = 11
    for x in range(10, 25):
        ws.row_dimensions[x].height = 22
    ws.row_dimensions[11].height = 35
    for row in rows:
        for cell in column:
            ws.cell(row=row, column=cell).border = thin_border

def colorCelda(ws):
    for x in abc('A', 'Q'):
        ws['{}10'.format(x)].fill = titleFill
        ws['{}23'.format(x)].fill = titleFill
        ws['{}24'.format(x)].fill = titleFill

def diasCuerpo(ws):
    n = 0
    lista = abc('B', 'P')
    for x in lista:
        n = n + 1
        ws["{}10".format(x)] = u'{}'.format(n)
        ws["{}10".format(x)].alignment = Alignment(horizontal="center", vertical="center")

def formatocampos(ws):
    m = 0
    n = 0
    lista1 = abc('B', 'P')
    lista2 = list(range(11,20))
    for p in lista1:
        m = m + 1
        for q in lista2:
            n = n + 1
            ws["{}{}".format(p, q)].number_format = "@"
        ws["{}20".format(p)].number_format = "@"


def cabecera(listubicacion, lista, wb, tipo, empadronadores):

    idemp = lista[1]
    if tipo == 1:
        scr = lista[2]
    else:
        scr = lista[-3:]
    ccdd = listubicacion[0]
    ccpp = listubicacion[1]
    ccdi = listubicacion[2]
    departamento = listubicacion[3]
    provincia = listubicacion[4]
    distrito = listubicacion[5]

    if tipo == 1:
        ws = wb.get_sheet_by_name(idemp)
    else:
        ws = wb.get_sheet_by_name(lista)
    ws["D1"] = u'CENSOS NACIONALES: XII DE LA POBLACION, VII DE VIVIENDA Y III DE COMUNIDADES INDIGENAS'
    ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
    ws["D1"].font = Font(bold=True)
    ws.merge_cells('D1:N1')
    ws["D2"] = u'III Censo de Comunidades Nativas y I Censo de Comunidades Campesinas'
    ws['D2'].alignment = Alignment(horizontal="center", vertical="center")
    ws["D2"].font = Font(bold=True)
    ws.merge_cells('D2:N2')

    ws["D4"] = u'PROGRAMACION DE RUTAS DE TRABAJO RURAL'
    ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
    ws["D4"].font = Font(bold=True, size=12)
    ws.merge_cells('D4:N4')

    ws["A6"] = u'DEPARTAMENTO'
    ws["A7"] = u'PROVINCIA'
    ws["A8"] = u'DISTRITO'
    ws["M6"] = u'EMPADRONADOR'
    ws["M8"] = u'JEFE DE SECCION'
    ws["A10"] = u'DIA'
    ws["A11"] = u'AER'
    ws["A12"] = u'ACTIVIDAD'
    ws["A23"] = u'G.OPER.'
    ws["A24"] = u'PASAJES'

    ws["A25"] = u'V: VIAJE'
    ws["A26"] = u'E: EMPADRONAMIENTO'
    ws["A27"] = u'T: TRASLADO'
    ws["A28"] = u'S: SUPERVISION'
    ws["Q5"] = u'Doc.CPV.03.105'
    ws["A25"].font = Font(bold=True)
    ws["A26"].font = Font(bold=True)
    ws["A27"].font = Font(bold=True)
    ws["A28"].font = Font(bold=True)
    ws["Q5"].font = Font(bold=True)
    ws["Q5"].alignment = Alignment(horizontal="right", vertical="bottom")

    ws["Q10"] = u'TOTAL'

    diasCuerpo(ws)
    for x in ['A10', 'A11', 'A12', 'A21', 'A22', 'Q8', 'Q10']:
        ws[x].alignment = Alignment(horizontal="center", vertical="center")

    colorCelda(ws)
    formatocampos(ws)

    infoubicacion = [ccdd, ccpp, ccdi, departamento, provincia, distrito]
    for x in range(6, 9):
        ws["C{}".format(x)] = infoubicacion[x-6]
        ws["D{}".format(x)] = infoubicacion[x-3]

    if tipo == 1:
        emp = lista[0]
        ws['O6'] = emp
        ws.cell(row=6, column=17).border = thin_border
    elif tipo == 2:
        listaEmp = []
        for i in empadronadores:
            if i[0] == lista:
                listaEmp.append(i[1])
            else:
                pass
        ws["O6"] = '-'.join(listaEmp)
        ws.cell(row=6, column=17).border = thin_border
        del listaEmp

    ws['O8'] = scr
    ws.cell(row=8, column=17).border = thin_border

    for x in range(13, 23):
        ws["A{}".format(x)] = u'CC.PP.'
        ws["A{}".format(x)].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('Q10:Q12')
    listaceldas = ['A6:B6', 'A7:B7', 'A8:B8', 'D6:H6', 'D7:H7', 'D8:H8'] # 'N6:O6', 'N8:O8', 'P6:Q6', 'P8:Q8'

    # for x in ['M6', 'M8', 'O6', 'O8', 'C6', 'C7', 'C8']:
    #     ws[x].alignment = Alignment(horizontal="center", vertical="center")

    for cells in listaceldas:
        ws.merge_cells(cells)
        set_border(cells, ws)

    for cells in [6, 7, 8]:
        ws.cell(row=cells, column=3).border = thin_border
    for cells in ['D6:H6', 'D7:H7', 'D8:H8']:
        ws.merge_cells(cells)
        set_border(cells, ws)
    cuerpo_border('A10:Q24', ws)

def excel2PDFemp(ubigeo, workspace):
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False
    wb_path = os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Empad.xlsx")
    empadronadores = infoEmpadronador(ubigeo)
    emp = []
    for m in empadronadores:
        emp.append(m[1])
    for i in emp:
        wb = o.Workbooks.Open(wb_path)
        wb.Sheets(wb.Worksheets[i].name).ExportAsFixedFormat(0, os.path.join(workspace, 'Rural', ubigeo, wb.Worksheets[i].name + "_Empad.pdf"))
        wb.Close()

def excel2PDFscr(ubigeo, workspace):
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False
    wb_path = os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_JSeccion.xlsx")
    jefeSeccion = infoSCR(ubigeo)
    for i in range(len(jefeSeccion)):
        wb = o.Workbooks.Open(wb_path)
        wb.Sheets(wb.Worksheets[i].name).ExportAsFixedFormat(0, os.path.join(workspace, 'Rural', ubigeo, wb.Worksheets[i].name + "_SCR.pdf"))
        wb.Close()

def generarExcelRutas(ubigeo, workspace):
    cabecera01 = infoGeneral(ubigeo)
    empadronadores = infoEmpadronador(ubigeo)
    wb = xlsxwriter.Workbook(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Empad.xlsx"))
    for x in empadronadores:
        ws = wb.add_worksheet(str(x[1]))
        ws.set_paper(9)
        ws.print_area(0, 0, 27, 16)
        ws.fit_to_pages(1, 1)
        ws.set_landscape()
    wb.close()
    wb = load_workbook(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Empad.xlsx"))
    for x in empadronadores:
        cabecera(cabecera01, x, wb, 1, empadronadores)
    wb.save(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Empad.xlsx"))

def generarExcelSCR(ubigeo, workspace):
    cabecera01 = infoGeneral(ubigeo)
    empadronadores = infoEmpadronadorSCR(ubigeo)
    jefeSeccion = infoSCR(ubigeo)
    wb = xlsxwriter.Workbook(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_JSeccion.xlsx"))
    for y in jefeSeccion:
        ws = wb.add_worksheet(str(y))
        ws.set_paper(9)
        ws.print_area(0, 0, 27, 16)
        ws.fit_to_pages(1, 1)
        ws.set_landscape()
    wb.close()
    wb = load_workbook(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_JSeccion.xlsx"))
    for y in jefeSeccion:
        cabecera(cabecera01, y, wb, 2, empadronadores)
    wb.save(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_JSeccion.xlsx"))


# Tipo 1: Empadronador
# Tipo 2: Jefe de Seccion