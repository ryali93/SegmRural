# -*-coding: utf-8-*-

from Generales import settings
from datetime import datetime
from openpyxl import load_workbook, drawing
from openpyxl.styles import Font
import string
import os
from ListadoProgramacionRutas2 import infoEmpadronador, infoSCR, set_border, generarExcelRutas, generarExcelSCR, Alignment, thin_border, actualizaIDEMP, actualizarAsignRural, excel2PDFemp, excel2PDFscr

startTime = datetime.now()
print startTime

conn = settings.conectionDB_pymmsql()
# conn2 = settings.conectionGDB_pymmsql()

def actualizaRutas(ubigeo):
    cursorsql2 = conn.cursor()
    cursorsql2.execute("EXEC ACTUALIZAR_RUTAS '{}'".format(ubigeo))
    conn.commit()
    cursorsql2.close()

def pasajeRural(id):
    informacion = []

    if len(id) == 20:
        cursor = conn.cursor()
        cursor.execute("SELECT COSTO_EMP FROM TB_PEA_RURAL_EMP WHERE IDEMP = '{}'".format(id))
        for el in cursor:
            informacion.append(el[0])
        cursor.close()

    elif len(id) == 14:
        cursor = conn.cursor()
        cursor.execute("SELECT COSTO_SCR FROM TB_PEA_RURAL_SCR WHERE IDSCR = '{}'".format(id))
        for el in cursor:
            informacion.append(el[0])
        cursor.close()
    return informacion[0]


def IdSCRUbigeo(ubigeo):
    cursor = conn.cursor()
    cursor.execute("SELECT IDSCR FROM SEGM_R_SCR WHERE UBIGEO = '{}' ORDER BY IDSCR".format(ubigeo))
    informacion = [x[0] for x in cursor]
    cursor.close()
    return informacion

def actualizarAsign(ubigeo):
    cursor = conn.cursor()
    cursor.execute("EXEC USP_ACTUALIZA_ASIGN_RURAL '{}'".format(ubigeo))
    conn.commit()
    cursor.close()

# Recolecta los campos de cada Ubigeo
def listar_ccpp_rural(idruta):
    registrosRuta = []
    cursor = conn.cursor()
    cursor.execute("SELECT UBIGEO, IDRUTA, DIA_RUTA, CODCCPP_RUTA, TIPO_RUTA, COSTO_RUTA FROM CPV_SEGMENTACION_GDB.SDE.SEGM_R_ASIGN_RUTA a WHERE IDRUTA='{}' ORDER BY DIA_RUTA, OBJECTID".format(idruta))
    for row in cursor.fetchall():
        registrosRuta.append(row)
    cursor.close()
    return registrosRuta

def listar_ccpp_rural_SCR(idruta):
    registrosSCR = []
    cursor = conn.cursor()
    cursor.execute("SELECT UBIGEO, IDSCR, DIA_SCR, CODCCPP_SCR, TIPO_SCR, COSTO_SCR FROM CPV_SEGMENTACION_GDB.SDE.SEGM_R_ASIGN_SCR a WHERE IDSCR='{}' ORDER BY DIA_SCR, OBJECTID".format(idruta))
    for row in cursor.fetchall():
        registrosSCR.append(row)
    cursor.close()
    return registrosSCR

def listarAER(ubigeo):
    listaAer = []
    cursor = conn.cursor()
    cursor.execute("SELECT CODCCPP, AER_INI, AER_FIN FROM CPV_SEGMENTACION_GDB.SDE.TB_CCPP WHERE UBIGEO = '{}' AND AREA = '2'".format(ubigeo))
    for row in cursor.fetchall():
        listaAer.append(row)
    cursor.close()
    return listaAer


def abc(inicial, final):
    abc = list(string.ascii_uppercase)
    ini = abc.index(inicial)
    fin = abc.index(final)
    lista = abc[ini:fin + 1]
    return lista

## Dar la forma para exportar a excel       x: IDRUTA
def formaExcel(x):
    lista_final = []
    while len(x) > 0:
        fila = []
        for i in range(15):
            encontro = 0
            if len(x) > 0:
                for ruta_ccpp in x:
                    if ruta_ccpp[2] == (i+1):
                        fila.append(ruta_ccpp[3])
                        x.remove(ruta_ccpp)
                        encontro = 1
                        break
                if encontro == 0:
                    fila.append('')
            else:
                fila.append('')
        lista_final.append(fila)
    return lista_final


## Llenar Excel
def llenarTabla(lista_final,ws):
    for fila in range(len(lista_final)):
        rangoLetra = abc("B", "P")
        for letra in abc("B", "P"):
            col = rangoLetra.index(letra)
            ws["{}{}".format(letra, fila+13)] = lista_final[fila][col]

def llenarInfo(x, ws):
    dicc = {}
    for dia in range(1, 16):
        dicc["dia_{}".format(dia)] = []

    #Dentro del for
    for col in abc("B", "P"):
        for i in x:
            if int(i[2]) == int(ws["{}10".format(col)].value):
                ws["{}12".format(col)] = i[4]
                # dicc["dia_{}".format(i[2])].append(i[5])
                # ws["{}24".format(col)] = sum(filter(None, dicc["dia_{}".format(i[2])]))
                ws["{}12".format(col)].alignment = Alignment(horizontal="center", vertical="center")
                # ws["{}24".format(col)].alignment = Alignment(horizontal="center", vertical="center")

        if ws["{}12".format(col)].value in ("S", "V", "E", "T"):
            ws["{}23".format(col)] = 30
    values = []
    # values2 = []
    for cell in abc("B", "P"):
        if ws["{}23".format(cell)].value is None:
            pass
        else:
            values.append(ws["{}23".format(cell)].value)
        # if ws["{}24".format(cell)].value == 0:
        #     ws["{}24".format(cell)] = None
        # else:
        #     values2.append(ws["{}24".format(cell)].value)
    ws["Q23"].value = sum(values)
    # ws["Q24"] = sum(filter(None, values2))
    # ws["Q24"] = sum(filter(None, values2))
    del values
    # del values2

def insertImage(ws):
    img = drawing.image.Image('D:/SegmentacionRuralV2/Insumos/Img/ENP_BW.png')
    img.drawing.width = 70
    img.drawing.height = 80
    ws.add_image(img, "B1")
    img2 = drawing.image.Image('D:/SegmentacionRuralV2/Insumos/Img/Inei_BW.png')
    img2.drawing.width = 100
    img2.drawing.height = 70
    ws.add_image(img2, "P1")

def agregarAer(ubigeo, ws):
    aers = listarAER(ubigeo)
    for col in abc("B", "P"):
        lista = []
        cantccpp = [ws['{}{}'.format(col, x)].value for x in range(13, 23)]
        if len(cantccpp) == 0:
            pass
        else:
            for i in cantccpp:
                for aer in aers:
                    if aer[0] == i:
                        if aer[1] == aer[2]:
                            lista.append(str(aer[1]))
                            # ws['{}11'.format(col)] = str(aer[1])
                        else:
                            lista.append(str(aer[1]) + "-" + str(aer[2]))
                            # ws['{}11'.format(col)] = str(aer[1]) + "-" + str(aer[2])
                lista = list(set(lista))
                cadena = "/ ".join(lista)
                ws['{}11'.format(col)] = cadena
        del lista
        for cells in range(11, 24):
            ws["{}{}".format(col, cells)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def tablasProgramacionRutas(ubigeo, workspace):
    generarExcelRutas(ubigeo, workspace)
    listEmp = []
    listRuta = []
    empadronadores = infoEmpadronador(ubigeo)
    for i in empadronadores:
        idemp = i[1]
        idruta = i[3]
        listEmp.append(idemp)
        listRuta.append(idruta)
    wb = load_workbook(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_Empad.xlsx"))
    for emp in listEmp:
        ws = wb.get_sheet_by_name(emp)
        for NomRuta in listRuta:
            if NomRuta == emp[:-3]:
                lista_ccpp_ruta = listar_ccpp_rural(NomRuta)
                x = lista_ccpp_ruta[:]
                llenarInfo(x, ws)
                listaFinal = formaExcel(lista_ccpp_ruta)
                llenarTabla(listaFinal, ws)
                insertImage(ws)
                for cell_f in range(10, 25):
                    ws["Q{}".format(cell_f)].alignment = Alignment(horizontal="center", vertical="center")
                    ws["Q{}".format(cell_f)].font = Font(bold=True)
                ws["Q10"].border = thin_border
                ws["Q11"].border = thin_border
                ws["Q12"].border = thin_border
            agregarAer(ubigeo, ws)
        ws.merge_cells('M6:N6')
        ws.merge_cells('M8:N8')
        ws.merge_cells('O6:Q6')
        ws.merge_cells('O8:Q8')
        for cells in ['A6:B6', 'A7:B7', 'A8:B8', 'D6:H6', 'D7:H7', 'D8:H8', 'M6:N6', 'M8:N8', 'O6:Q6', 'O8:Q8']:
            set_border(cells, ws)
        for x in ['M6', 'M8', 'O6', 'O8', 'C6', 'C7', 'C8']:
            ws[x].alignment = Alignment(horizontal="center", vertical="center")
        pasaje = pasajeRural(emp)
        ws['Q24'] = pasaje

    wb.save(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_Empad.xlsx"))


def tablasProgramacionSCR(ubigeo, workspace):
    generarExcelSCR(ubigeo, workspace)
    jefeSeccion = infoSCR(ubigeo)
    wb = load_workbook(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_JSeccion.xlsx"))
    for jf in jefeSeccion:
        ws = wb.get_sheet_by_name(jf)
        lista_ccpp_SCR = listar_ccpp_rural_SCR(jf)
        x = lista_ccpp_SCR[:]
        llenarInfo(x, ws)
        listaFinal = formaExcel(x)
        llenarTabla(listaFinal, ws)
        agregarAer(ubigeo, ws)
        ws.merge_cells('M6:N6')
        ws.merge_cells('M8:N8')
        ws.merge_cells('O6:Q6')
        ws.merge_cells('O8:Q8')
        for cells in ['A6:B6', 'A7:B7', 'A8:B8', 'D6:H6', 'D7:H7', 'D8:H8', 'M6:N6', 'M8:N8', 'O6:Q6', 'O8:Q8']:
            set_border(cells, ws)
        insertImage(ws)
        for cell_f in range(10, 25):
            ws["Q{}".format(cell_f)].alignment = Alignment(horizontal="center", vertical="center")
            ws["Q{}".format(cell_f)].font = Font(bold=True)
        ws["Q10"].border = thin_border
        ws["Q11"].border = thin_border
        ws["Q12"].border = thin_border
        for x in ['M6', 'M8', 'O6', 'O8', 'C6', 'C7', 'C8']:
            ws[x].alignment = Alignment(horizontal="center", vertical="center")
        pasaje = pasajeRural(jf)
        if pasaje == 0:
            letras = abc('B', 'Q')
            for letra in letras:
                ws['{}23'.format(letra)] = 0
        ws['Q24'] = pasaje

    wb.save(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_JSeccion.xlsx"))

def programacionRutas(ubigeo, workspace):
    actualizaRutas(ubigeo)
    actualizaIDEMP(ubigeo)
    actualizarAsignRural(ubigeo)
    generarExcelRutas(ubigeo, workspace)
    generarExcelSCR(ubigeo, workspace)
    tablasProgramacionRutas(ubigeo, workspace)
    tablasProgramacionSCR(ubigeo, workspace)
    excel2PDFemp(ubigeo, workspace)
    excel2PDFscr(ubigeo, workspace)
    print ubigeo

# ubigeo = '010706'
# workspace = r'D:\SegmentacionRuralV2\Procesamiento\ProgRutas'
# programacionRutas(ubigeo, workspace)