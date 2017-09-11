# -*-coding: utf-8-*-

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl import drawing
from Generales import settings
import win32com.client
import os
import string
import xlsxwriter

conn = settings.conectionDB_pymmsql()

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

titleFill = PatternFill(start_color='D9D9D9',
                        end_color='D9D9D9',
                        fill_type='solid')

def pasajeRural(id):
    informacion = []
    if len(id) == 20:
        cursor = conn.cursor()
        cursor.execute("SELECT COSTO_EMP FROM TB_PEA_RURAL_EMP WHERE IDEMP = '{}'".format(id))
        for el in cursor:
            informacion.append(el[0])
        cursor.close()
    elif len(id) == 14:
        cursor = conn.cursor()
        cursor.execute("SELECT COSTO_SCR FROM TB_PEA_RURAL_SCR WHERE IDSCR = '{}'".format(id))
        for el in cursor:
            informacion.append(el[0])
        cursor.close()
    return informacion[0]


def abc(inicial, final):
    abc = list(string.ascii_uppercase)
    ini = abc.index(inicial)
    fin = abc.index(final)
    lista = abc[ini:fin + 1]
    return lista

def infoEmpadronadorResumen(ubigeo, conn=conn):
    registrosResEmp = []
    cursor = conn.cursor()
    expresion_sql = "SELECT R.SCR, R.RUTA, E.EMP, R.VIAJE, R.TRASLADO, R.EMPADRONAMIENTO, R.SUPERVISION, R.DIA_TOTAL AS TOTAL, R.PASAJE, R.DIA_TOTAL*30 AS VIATICOS, (R.PASAJE + R.DIA_TOTAL*30) AS TOTAL_GENERAL, R.NUM_CCPP, F.N_VIV_RUTA FROM (SELECT SUBSTRING(IDRUTA,12,3) AS SCR, SUBSTRING(IDRUTA,15,3) AS RUTA, IDRUTA, COUNT(CASE WHEN TIPO_RUTA = 'V' then 1 else null end) AS VIAJE, COUNT(CASE WHEN TIPO_RUTA = 'T' then 1 else null end) AS TRASLADO, (SELECT COUNT(DISTINCT DIA_RUTA) FROM CPV_SEGMENTACION_GDB.SDE.SEGM_R_ASIGN_RUTA  X WHERE  X.IDRUTA=xx.IDRUTA AND TIPO_RUTA='E' )  AS  EMPADRONAMIENTO, '0' AS SUPERVISION, MAX(DIA_RUTA) AS DIA_TOTAL,SUM(COSTO_RUTA) AS PASAJE, COUNT(*) AS VIATICOS, COUNT(DISTINCT CODCCPP) AS NUM_CCPP FROM CPV_SEGMENTACION_GDB.SDE.SEGM_R_ASIGN_RUTA xx WHERE UBIGEO = '{}' GROUP BY IDRUTA) AS R INNER JOIN (SELECT IDRUTA, EMP FROM SEGM_R_EMP) E ON R.IDRUTA COLLATE DATABASE_DEFAULT = E.IDRUTA COLLATE DATABASE_DEFAULT INNER JOIN (SELECT IDRUTA, N_VIV_RUTA, N_EMP_RUTA FROM CPV_SEGMENTACION_GDB.SDE.SEGM_R_RUTA) F ON R.IDRUTA COLLATE DATABASE_DEFAULT = F.IDRUTA COLLATE DATABASE_DEFAULT ORDER BY RUTA".format(ubigeo)
    cursor.execute(expresion_sql)
    for row in cursor.fetchall():
        registrosResEmp.append(row)
    cursor.close()
    return registrosResEmp

def infoSCRresumen(ubigeo, conn=conn):
    registrosResSCR = []
    cursor = conn.cursor()
    expresion_sql = "SELECT S.SCR, '' AS RUTA, '' AS EMP, S.VIAJE, S.TRASLADO, S.EMPADRONAMIENTO, S.SUPERVISION, S.DIA_TOTAL AS TOTAL, S.PASAJE, S.DIA_TOTAL*30 AS VIATICOS, (S.PASAJE + S.DIA_TOTAL*30) AS TOTAL_GENERAL,  '' AS NUM_CCPP, '' AS N_VIV_RUTA FROM (SELECT SUBSTRING(IDSCR,12,3) AS SCR, COUNT(CASE WHEN TIPO_SCR = 'V' then 1 else null end) AS VIAJE, COUNT(CASE WHEN TIPO_SCR = 'T' then 1 else null end) AS TRASLADO, '0' AS EMPADRONAMIENTO, (SELECT COUNT(DISTINCT DIA_SCR) FROM CPV_SEGMENTACION_GDB.SDE.SEGM_R_ASIGN_SCR  X WHERE  X.IDSCR=xx.IDSCR AND TIPO_SCR='S' )  AS SUPERVISION, MAX(DIA_SCR) AS DIA_TOTAL, SUM(COSTO_SCR) AS PASAJE, COUNT(*) AS VIATICOS, COUNT(CODCCPP) AS NUM_CCPP FROM CPV_SEGMENTACION_GDB.SDE.SEGM_R_ASIGN_SCR XX WHERE UBIGEO = '{}' GROUP BY IDSCR) AS S ORDER BY SCR".format(ubigeo)
    cursor.execute(expresion_sql)
    for row in cursor.fetchall():
        registrosResSCR.append(row)
    cursor.close()
    return registrosResSCR

def infoSCR(ubigeo, conn=conn):
    cursor = conn.cursor()
    cursor.execute("SELECT SCR FROM SEGM_R_SCR WHERE UBIGEO = '{}' ORDER BY SCR".format(ubigeo))
    informacion = [x[0] for x in cursor]
    cursor.close()
    return informacion

def infoGeneral(ubigeo, conn=conn):
    cursor = conn.cursor()
    cursor.execute("SELECT DEPARTAMENTO, PROVINCIA, DISTRITO FROM CPV_SEGMENTACION.dbo.VW_DISTRITO WHERE UBIGEO = '{}'".format(ubigeo))
    informacion = [[ubigeo[0:2], ubigeo[2:4], ubigeo[4:6], x[0], x[1], x[2]] for x in cursor]
    cursor.close()
    return informacion[0]

def set_border(ws):

    list_col = abc('B', 'E') + abc('K', 'O')
    for i in list_col:
        ws["{}11".format(i)].border = thin_border
        ws["{}12".format(i)].border = thin_border
    list_fil = abc('E','G')
    for i in list_fil:
        ws["{}6".format(i)].border = thin_border
        ws["{}7".format(i)].border = thin_border
        ws["{}8".format(i)].border = thin_border
    list_fil2 = abc('F', 'M')
    for i in list_fil2:
        ws["{}10".format(i)].border = thin_border


def insertImage(ws):
    img = drawing.image.Image('D:/SegmentacionRuralV2/Insumos/Img/ENP_BW.png')
    img.drawing.width = 70
    img.drawing.height = 80
    ws.add_image(img, "C1")
    img2 = drawing.image.Image('D:/SegmentacionRuralV2/Insumos/Img/Inei_BW.png')
    img2.drawing.width = 100
    img2.drawing.height = 70
    ws.add_image(img2, "N1")

def cuerpo_border(ws, alto):
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 8
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 2

    ws.row_dimensions[10].height = 20
    ws.row_dimensions[11].height = 20
    ws.row_dimensions[12].height = 20

    for row in range(10) + range(13,26):
        ws.row_dimensions[row].height = 15

    columns = range(2,16)
    rows = range(13, 14+alto)
    for col in columns:
        for row in rows:
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

def colorCelda(ws):
    ws['B13'].fill = titleFill
    for x in abc('B', 'O'):
        ws['{}10'.format(x)].fill = titleFill
        ws['{}11'.format(x)].fill = titleFill
        ws['{}12'.format(x)].fill = titleFill

def diasCuerpo(ws):
    n = 0
    lista = abc('B', 'P')
    for x in lista:
        n = n + 1
        ws["{}10".format(x)] = u'{}'.format(n)
        ws["{}10".format(x)].alignment = Alignment(horizontal="center", vertical="center")

def formatocampos(ws):
    m = 0
    n = 0
    lista1 = abc('B', 'E')
    lista2 = list(range(14, 50))
    for p in lista1:
        m = m + 1
        for q in lista2:
            n = n + 1
            ws["{}{}".format(p, q)].number_format = "@"

def llenarTotal(ws):
    for x in abc('F', 'O'):
        ws['{}13'.format(x)] = "=SUM({}14: {}700)".format(x, x)

def cabecera(listubicacion, wb, ubigeo):

    ccdd = listubicacion[0]
    ccpp = listubicacion[1]
    ccdi = listubicacion[2]
    departamento = listubicacion[3]
    provincia = listubicacion[4]
    distrito = listubicacion[5]

    ws = wb.get_sheet_by_name(ubigeo)

    ws["D1"] = u'CENSOS NACIONALES: XII DE LA POBLACION, VII DE VIVIENDAY III DE COMUNIDADES INDIGENAS'
    ws["D2"] = u'III Censo de Comunidades Nativas y I Censo de Comunidades Campesinas'
    ws["B4"] = u'RESUMEN DISTRITAL DE LA PROGRAMACION DE RUTAS DEL AREA RURAL'
    ws["B4"].font = Font(size=14)

    ws["B6"] = u'DEPARTAMENTO'
    ws["B7"] = u'PROVINCIA'
    ws["B8"] = u'DISTRITO'

    ws["N6"] = u'Doc.CPV.03.105A'
    ws["N6"].alignment = Alignment(horizontal="right", vertical="bottom")
    ws["N6"].font = Font(bold=True)
    ws["B10"] = u'N° ORD'
    ws["C10"] = u'JEFE/A DE SECCIÓN N°'
    ws["D10"] = u'RUTA N°'
    ws["E10"] = u'EMP N°'
    ws["F10"] = u'DIAS DE OPERACION DE CAMPO'
    ws["K10"] = u'ASIGNACION DE FONDOS'
    ws["N10"] = u'NÚMERO DE CCPP'
    ws["O10"] = u'NÚMERO DE VIVIENDAS'
    ws["F12"] = u'VIAJE'
    ws["G12"] = u'TRASLADO'
    ws["H12"] = u'EMPADRONADORES'
    ws["I12"] = u'SUPERVISIÓN'
    ws["J12"] = u'TOTAL'
    ws["K11"] = u'PASAJE (S/.)'
    ws["L11"] = u'VIÁTICOS (S/.30)'
    ws["M11"] = u'TOTAL GENERAL (S/.)'
    ws["B13"] = u'TOTAL'

    celdasCabecera = ['E6:G6','E7:G7','E8:G8','D1:M1','D2:M2','B4:O4','B10:B12', 'C10:C12', 'D10:D12', 'E10:E12', 'F10:J11', 'K10:M10','N10:N12', 'O10:O12', 'K11:K12', 'L11:L12', 'M11:M12', 'N6:O6', 'B13:E13']
    for cells in celdasCabecera:
        ws.merge_cells(cells)

    listaceldas2 = ['D1', 'D2', 'B4', 'O6']
    for cells in listaceldas2:
        ws[cells].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cells  == 'B4':
            ws[cells].font = Font(bold=True, size=14)
        else:
            ws[cells].font = Font(bold=True, size=12)

    listaceldas3 = ['B10','C10','D10','E10','F10','K10','N10','O10','F12','G12','H12','I12','J12','K11','L11','M11','B13']
    for cells in listaceldas3:
        ws[cells].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[cells].font = Font(bold=True, size=8)

    listaceldas4 = []
    rows = range(10, 13)
    column = abc('B', 'E') + ['N','O']
    rows2 = range(11, 13)
    column2 = abc('K', 'M')

    for row in rows:
        for cell in column:
            listaceldas4.append('{}{}'.format(cell, row))
    for row in rows2:
        for cell in column2:
            listaceldas4.append('{}{}'.format(cell, row))
    for i in ('F10', 'G10', 'H10', 'I10', 'J10', 'F12', 'G12', 'H12', 'I12', 'J12', 'K10', 'L10', 'M10','B13'):
        listaceldas4.append(i)
    for cells in listaceldas4:
        ws[cells].border = thin_border
    for rows in [2, 3, 4, 5, 6, 7]:
        for cols in [6, 7, 8]:
            if cols == 6:
                ws.cell(row=cols, column=rows).font = Font(bold=True)
                ws.cell(row=cols, column=rows).font = Font(bold=True)
            else:
                pass
            ws.cell(row=cols, column=rows).border = thin_border
    colorCelda(ws)

    infoubicacion = [ccdd, ccpp, ccdi, departamento, provincia, distrito]
    for x in range(6, 9):
        ws["D{}".format(x)] = infoubicacion[x-6]
        ws["E{}".format(x)] = infoubicacion[x-3]


def generarResumenRutas(ubigeo, workspace, alto):
    # alto1 = len(infoEmpadronadorResumen(ubigeo))
    # alto2 = len(infoSCRresumen(ubigeo))
    # alto = alto1 + alto2
    # empadronadores = infoEmpadronador(ubigeo)
    wb = xlsxwriter.Workbook(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Resumen.xlsx"))
    ws = wb.add_worksheet('{}'.format(ubigeo))
    ws.print_area(1, 1, alto+12, 14)
    # print alto
    ws.fit_to_pages(1, 1)
    wb.close()

def excel2PDF(ubigeo, workspace):
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False
    wb_path = os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Resumen.xlsx")
    wb = o.Workbooks.Open(wb_path)
    wb.Sheets(ubigeo).ExportAsFixedFormat(0, os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Resumen.pdf"))
    # wb.save(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Resumen.xlsx"))
    wb.Close(True)

def llenarTablaResumen(ubigeo, workspace, alto):
    cabecera01 = infoGeneral(ubigeo)
    infoEmp = infoEmpadronadorResumen(ubigeo)
    SCRcampos = infoSCRresumen(ubigeo)
    listaSCR = infoSCR(ubigeo)
    letras = abc('C', 'O')
    columnas = range(len(letras))
    print ubigeo
    wb = load_workbook(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Resumen.xlsx"))
    cabecera(cabecera01, wb, ubigeo)
    ws = wb.get_sheet_by_name(ubigeo)
    formatocampos(ws)

    fila = 13

    for i in listaSCR:
        # print i
        for SCR in SCRcampos:
            if SCR[0] == i:
                # print "SCR: " + str(ubigeo) + "00000" + str(SCR[0])
                idscr = str(ubigeo) + "00000" + str(SCR[0])
                pasaje = pasajeRural(idscr)
                fila += 1
                for col in columnas:
                    ws["{}{}".format(letras[col], fila)] = SCR[col]
                if pasaje == 0:
                    ws['L{}'.format(fila)] = 0
                ws["K{}".format(fila)] = pasaje
                ws["M{}".format(fila)] = pasaje + ws["L{}".format(fila)].value
                    # print str(letras[col]) + "-" + str(fila) + " : " + str(SCR[col])
            for empCampos in infoEmp:
                if empCampos[0] == SCR[0] and SCR[0] == i:
                    # print "    Emp: " + str(ubigeo) + "00000" + str(SCR[0]) + str(empCampos[1]) + str(empCampos[2])
                    idemp = str(ubigeo) + "00000" + str(SCR[0]) + str(empCampos[1]) + str(empCampos[2])
                    pasaje = pasajeRural(idemp)
                    fila += 1
                    for col in columnas:
                        ws["{}{}".format(letras[col], fila)] = empCampos[col]
                    ws["K{}".format(fila)] = pasaje
                    ws["M{}".format(fila)] = pasaje + ws["L{}".format(fila)].value
                    # print str(letras[col]) + "-" + str(fila) + " : " + str(empCampos[col])
    for i in range(14, 14+alto):
        ws["B{}".format(i)] = i-13

    insertImage(ws)
    cuerpo_border(ws, alto)
    for i in range(14, 200):
        if ws["D{}".format(i)].value == ws["D{}".format(i-1)].value:
            ws["N{}".format(i)] = ""
            ws["O{}".format(i)] = ""
    llenarTotal(ws)
    set_border(ws)
    # ws = wb.active
    # ws.print_area = 'B1:O{}'.format(13+alto)
    # ws.print_area(1, 1, alto + 12, 14)
    wb.save(os.path.join(workspace, 'Rural', ubigeo, ubigeo + "_Resumen.xlsx"))

def tablaResumen(ubigeo, workspace):
    alto1 = len(infoEmpadronadorResumen(ubigeo))
    alto2 = len(infoSCRresumen(ubigeo))
    alto = alto1 + alto2
    generarResumenRutas(ubigeo, workspace, alto)
    llenarTablaResumen(ubigeo, workspace, alto)
    excel2PDF(ubigeo, workspace)


# ubigeo = '060701'
# workspace = "D:\SegmentacionRuralV2\Procesamiento\ProgRutas"
# tablaResumen(ubigeo, workspace)

