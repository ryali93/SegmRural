# -*-coding: utf-8-*-

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl import drawing
from Generales import settings
import win32com.client
import os
import string
import xlsxwriter

conn = settings.conectionDB_pymmsql()

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

titleFill = PatternFill(start_color='D9D9D9',
                        end_color='D9D9D9',
                        fill_type='solid')

def abc(inicial, final):
    abc = list(string.ascii_uppercase)
    ini = abc.index(inicial)
    fin = abc.index(final)
    lista = abc[ini:fin + 1]
    return lista

def infoEmpadronador(ubigeo, conn=conn):
    cursor = conn.cursor()
    cursor.execute("SELECT EMP, IDEMP, SCR, IDRUTA FROM CPV_SEGMENTACION.dbo.SEGM_R_EMP WHERE UBIGEO = '{}' ORDER BY EMP".format(ubigeo))
    informacion = [[x[0], x[1], x[2], x[3]] for x in cursor]
    cursor.close()
    return informacion


def infoSCR(ubigeo, conn=conn):
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT IDSCR FROM CPV_SEGMENTACION.dbo.SEGM_R_EMP WHERE UBIGEO = '{}'".format(ubigeo))
    informacion = [x[0] for x in cursor]
    cursor.close()
    return informacion


def infoGeneral(ubigeo, conn=conn):
    cursor = conn.cursor()
    cursor.execute("SELECT DEPARTAMENTO, PROVINCIA, DISTRITO FROM CPV_SEGMENTACION.dbo.VW_DISTRITO WHERE UBIGEO = '{}'".format(ubigeo))
    informacion = [[ubigeo[0:2], ubigeo[2:4], ubigeo[4:6], x[0], x[1], x[2]] for x in cursor]
    cursor.close()
    return informacion[0]

def set_border(cell_range, ws):
    lista = abc('A', 'Q')
    row = cell_range.split(":")[0][1]
    ini = cell_range.split(":")[0][0]
    fin = cell_range.split(":")[1][0]
    column = range(lista.index(ini) + 1, lista.index(fin) + 2)
    for cell in column:
        ws.cell(row=int(row), column=cell).border = thin_border

def cuerpo_border(cell_range, ws):
    lista = abc('A', 'Q')
    rowini = cell_range.split(":")[0][1:3]
    rowfin = cell_range.split(":")[1][1:3]
    ini = cell_range.split(":")[0][0]
    fin = cell_range.split(":")[1][0]
    rows = range(int(rowini), int(rowfin) + 1)
    column = range(lista.index(ini) + 1, lista.index(fin) + 2)
    for x in lista:
        ws.column_dimensions[x].width = 10
    for x in range(8, 23):
        ws.row_dimensions[x].height = 20
    for row in rows:
        for cell in column:
            ws.cell(row=row, column=cell).border = thin_border

def colorCelda(ws):
    for x in abc('A', 'Q'):
        ws['{}8'.format(x)].fill = titleFill
        # ws['{}18'.format(x)].fill = titleFill
        # ws['{}19'.format(x)].fill = titleFill
        ws['{}21'.format(x)].fill = titleFill
        ws['{}22'.format(x)].fill = titleFill

def diasCuerpo(ws):
    n = 0
    lista = abc('B', 'P')
    for x in lista:
        n = n + 1
        ws["{}8".format(x)] = u'{}'.format(n)
        ws["{}8".format(x)].alignment = Alignment(horizontal="center", vertical="center")

def formatocampos(ws):
    m = 0
    n = 0
    lista1 = abc('B', 'P')
    lista2 = list(range(9,18))
    for p in lista1:
        m = m + 1
        for q in lista2:
            n = n + 1
            ws["{}{}".format(p, q)].number_format = "@"
        ws["{}20".format(p)].number_format = "@"

def insertImage(ws):
    img = drawing.image.Image('D:/SegmentacionRuralV2/Insumos/Img/ENP_BW.png')
    img.drawing.width = 50
    img.drawing.height = 60
    img.anchor(ws["B2"])
    ws.add_image(img)
    img2 = drawing.image.Image('D:/SegmentacionRuralV2/Insumos/Img/Inei_BW.png')
    img2.drawing.width = 50
    img2.drawing.height = 50
    img2.anchor(ws["P2"])
    ws.add_image(img2)


def cabecera(listubicacion, lista, wb, tipo):
    emp = lista[0]
    idemp = lista[1]
    if tipo == 1:
        scr = lista[2]
    else:
        scr = lista[-3:]
    ccdd = listubicacion[0]
    ccpp = listubicacion[1]
    ccdi = listubicacion[2]
    departamento = listubicacion[3]
    provincia = listubicacion[4]
    distrito = listubicacion[5]

    if tipo == 1:
        ws = wb.get_sheet_by_name(idemp)
    else:
        ws = wb.get_sheet_by_name(lista)
    ws["A1"] = u'CENSOS NACIONALES: XII DE LA POBLACION, VII DE VIVIENDAY III DE COMUNIDADES INDIGENAS'
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True)
    ws.merge_cells('A1:Q1')
    ws["A2"] = u'PROGRAMACION DE RUTAS DE TRABAJO RURAL'
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = Font(bold=True)
    ws["A4"] = u'DEPARTAMENTO'
    ws["A5"] = u'PROVINCIA'
    ws["A6"] = u'DISTRITO'
    if tipo == 1:
        ws["N4"] = u'EMPADRONADOR'
    else:
        pass
    ws["N6"] = u'JEFE DE SECCION'
    ws["A8"] = u'DIA'
    ws["A9"] = u'AER'
    ws["A10"] = u'ACTIVIDAD'
    # ws["A18"] = u'G.OPER.'
    # ws["A19"] = u'PASAJES'
    # ws["A20"] = u'JEF.SECC.'
    ws["A21"] = u'G.OPER.'
    ws["A22"] = u'PASAJES'
    ws["Q8"] = u'TOTAL'

    diasCuerpo(ws)
    # for x in ['A8', 'A9', 'A10', 'A18', 'A19', 'A20', 'A21', 'A22', 'Q8']:
    for x in ['A8', 'A9', 'A10', 'A21', 'A22', 'Q8']:
        ws[x].alignment = Alignment(horizontal="center", vertical="center")

    colorCelda(ws)
    formatocampos(ws)
    insertImage(ws)

    infoubicacion = [ccdd, ccpp, ccdi, departamento, provincia, distrito]
    for x in range(4, 7):
        ws["C{}".format(x)] = infoubicacion[x-4]
        ws["D{}".format(x)] = infoubicacion[x-1]

    if tipo == 1:
        ws['Q4'] = emp
        ws.cell(row=4, column=17).border = thin_border
    else:
        pass
    ws['Q6'] = scr
    ws.cell(row=6, column=17).border = thin_border

    for x in ['Q4', 'Q6', 'C4', 'C5', 'C6']:
        ws[x].alignment = Alignment(horizontal="center", vertical="center")

    for x in range(11, 21):
        ws["A{}".format(x)] = u'CC.PP.'
        ws["A{}".format(x)].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A2:Q2')
    ws.merge_cells('Q8:Q10')
    if tipo == 1:
        listaceldas = ['A4:B4', 'A5:B5', 'A6:B6', 'D4:H4', 'D5:H5', 'D6:H6', 'N6:P6', 'N4:P4']
    else:
        listaceldas = ['A4:B4', 'A5:B5', 'A6:B6', 'D4:H4', 'D5:H5', 'D6:H6', 'N6:P6']
    for cells in listaceldas:
        ws.merge_cells(cells)
        set_border(cells, ws)

    for cells in [4, 5, 6]:
        ws.cell(row=cells, column=3).border = thin_border
    for cells in ['D4:H4', 'D5:H5', 'D6:H6']:
        ws.merge_cells(cells)
        set_border(cells, ws)
    cuerpo_border('A8:Q22', ws)


def excel2PDF(ubigeo, workspace):
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False
    wb_path = os.path.join(workspace, 'rural', ubigeo, ubigeo + ".xlsx")
    wb = o.Workbooks.Open(wb_path)
    path_to_pdf = os.path.join(workspace, 'rural', ubigeo, ubigeo + ".pdf")
    wb.ExportAsFixedFormat(0, path_to_pdf)

def generarExcelRutas(ubigeo, workspace):
    cabecera01 = infoGeneral(ubigeo)
    empadronadores = infoEmpadronador(ubigeo)
    wb = xlsxwriter.Workbook(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_Empad.xlsx"))
    for x in empadronadores:
        ws = wb.add_worksheet(str(x[1]))
        ws.print_area(0, 0, 21, 16)
        ws.fit_to_pages(1, 1)
        ws.set_landscape()
    wb.close()
    wb = load_workbook(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_Empad.xlsx"))
    for x in empadronadores:
        cabecera(cabecera01, x, wb, 1)
    wb.save(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_Empad.xlsx"))
    print ubigeo

def generarExcelSCR(ubigeo, workspace):
    cabecera01 = infoGeneral(ubigeo)
    jefeSeccion = infoSCR(ubigeo)
    wb = xlsxwriter.Workbook(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_JSeccion.xlsx"))
    for y in jefeSeccion:
        ws = wb.add_worksheet(str(y))
        ws.print_area(0, 0, 21, 16)
        ws.fit_to_pages(1, 1)
        ws.set_landscape()
    wb.close()
    wb = load_workbook(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_JSeccion.xlsx"))
    for y in jefeSeccion:
        cabecera(cabecera01, y, wb, 2)
    wb.save(os.path.join(workspace, 'rural', ubigeo, ubigeo + "_JSeccion.xlsx"))
    print ubigeo

# generarExcelRutas('150604', r'C:\Users\otin013\Desktop\PRUEBA')
# generarExcelSCR('150604', r'C:\Users\otin013\Desktop\PRUEBA')
# En Cabecera:
# Tipo 1: Empadronador
# Tipo 2: Jefe de Seccion